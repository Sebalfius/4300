import os
import wx
import sys
import json
import wx.adv
import wx.grid
import requests
import subprocess
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

if getattr(sys, 'frozen', False):
    # Cuando se ejecuta el programa como un .exe
    app_dir = os.path.dirname(sys.executable)
else:
    # Cuando se ejecuta el programa como un script .py
    app_dir = os.path.dirname(os.path.abspath(__file__))

credentials_path = os.path.join(app_dir, "credentialsforapi.json")
authkey_path = os.path.join(app_dir, "authkey.json")

class TextRedirect:
    """Redirect print statements to a wx.TextCtrl."""
    def __init__(self, text_ctrl):
        self.text_ctrl = text_ctrl

    def write(self, message):
        self.text_ctrl.AppendText(message)

    def flush(self):
        pass

class UpdateCredentialsDialog(wx.Dialog):
    """Dialog for updating the login credentials."""
    def __init__(self, parent):
        super().__init__(parent, title="Update Credentials", size=(350, 200))

        # Set up the dialog layout
        vbox = wx.BoxSizer(wx.VERTICAL)

        # Username field
        hbox_username = wx.BoxSizer(wx.HORIZONTAL)
        lbl_username = wx.StaticText(self, label="Username:")
        self.txt_username = wx.TextCtrl(self)
        hbox_username.Add(lbl_username, flag=wx.RIGHT, border=8)
        hbox_username.Add(self.txt_username, proportion=1)

        # Password field
        hbox_password = wx.BoxSizer(wx.HORIZONTAL)
        lbl_password = wx.StaticText(self, label="Password:")
        self.txt_password = wx.TextCtrl(self, style=wx.TE_PASSWORD)
        hbox_password.Add(lbl_password, flag=wx.RIGHT, border=8)
        hbox_password.Add(self.txt_password, proportion=1)

        # Load current values into the text fields
        self.load_credentials()

        # Buttons for Save and Cancel
        hbox_buttons = wx.BoxSizer(wx.HORIZONTAL)
        btn_save = wx.Button(self, label="Save")
        btn_cancel = wx.Button(self, label="Cancel")
        hbox_buttons.Add(btn_save)
        hbox_buttons.Add(btn_cancel, flag=wx.LEFT, border=10)

        # Bind button events
        btn_save.Bind(wx.EVT_BUTTON, self.on_save)
        btn_cancel.Bind(wx.EVT_BUTTON, self.on_cancel)

        # Add everything to the vbox
        vbox.Add(hbox_username, flag=wx.EXPAND | wx.ALL, border=10)
        vbox.Add(hbox_password, flag=wx.EXPAND | wx.ALL, border=10)
        vbox.Add(hbox_buttons, flag=wx.ALIGN_CENTER | wx.ALL, border=10)

        self.SetSizer(vbox)

    def load_credentials(self):
        """Load existing username and password from credentials.json file"""
        if os.path.exists(credentials_path):
            with open(credentials_path, 'r') as f:
                credentials = json.load(f)
                self.txt_username.SetValue(credentials.get("username", ""))
                self.txt_password.SetValue(credentials.get("password", ""))

    def on_save(self, event):
        """Save the new username and password to credentials.json file"""
        new_username = self.txt_username.GetValue()
        new_password = self.txt_password.GetValue()

        # Update JSON file
        with open(credentials_path, 'r+') as f:
            credentials = json.load(f)
            credentials["username"] = new_username
            credentials["password"] = new_password
            f.seek(0)
            json.dump(credentials, f, indent=4)
            f.truncate()  # Clear any remaining old data in the file after the new content

        wx.MessageBox("Credentials updated successfully.", "Success", wx.OK | wx.ICON_INFORMATION)
        self.EndModal(wx.ID_OK)

    def on_cancel(self, event):
        """Close the dialog without saving"""
        self.EndModal(wx.ID_CANCEL)

class AccountManager(wx.Frame):
    def __init__(self, parent):
        super().__init__(parent, title="Account Manager", size=(400, 300))
        self.panel = wx.Panel(self)
        self.sizer = wx.BoxSizer(wx.VERTICAL)
        
        # Load accounts from file
        self.accounts = self.load_accounts("listofaccountstoread.txt")
        
        # ListBox to display accounts
        self.account_list = wx.ListBox(self.panel, choices=self.accounts, style=wx.LB_SINGLE)
        self.sizer.Add(self.account_list, proportion=1, flag=wx.ALL | wx.EXPAND, border=10)

        # Buttons for operations
        self.add_button = wx.Button(self.panel, label="Add Account")
        self.remove_button = wx.Button(self.panel, label="Remove Selected")
        self.save_button = wx.Button(self.panel, label="Save Changes")
        #self.cancel_button = wx.Button(self.panel, label="Cancel")
        
        button_sizer = wx.BoxSizer(wx.HORIZONTAL)
        button_sizer.Add(self.add_button, flag=wx.RIGHT, border=5)
        button_sizer.Add(self.remove_button, flag=wx.RIGHT, border=5)
        button_sizer.Add(self.save_button, flag=wx.RIGHT, border=5)
        #button_sizer.Add(self.cancel_button, flag=wx.RIGHT, border=5)
        self.sizer.Add(button_sizer, flag=wx.ALL | wx.ALIGN_CENTER, border=10)

        # Bind events
        self.add_button.Bind(wx.EVT_BUTTON, self.on_add_account)
        self.remove_button.Bind(wx.EVT_BUTTON, self.on_remove_selected)
        self.save_button.Bind(wx.EVT_BUTTON, self.on_save_changes)
        #self.cancel_button.Bind(wx.EVT_BUTTON, self.on_cancelation)

        # Set sizer and center the window
        self.panel.SetSizer(self.sizer)
        self.Centre()
        self.Show()

    def load_accounts(self, filepath):
        """Load accounts from a file."""
        try:
            with open(filepath, "r") as f:
                return [line.strip() for line in f if line.strip()]
        except FileNotFoundError:
            return []

    def save_accounts(self, file_path):
        """Save accounts to the specified file."""
        with open(file_path, "w") as file:
            for account in self.accounts:
                file.write(account + "\n")    

    def on_add_account(self, event):
        """Add a new account."""
        dialog = wx.TextEntryDialog(self, "Enter a 4-digit account number:", "Add Account")
        if dialog.ShowModal() == wx.ID_OK:
            account = dialog.GetValue().strip()
            if account.isdigit() and len(account) == 4 and account not in self.accounts:
                self.accounts.append(account)
                self.account_list.Append(account)
            else:
                wx.MessageBox("Invalid account number. It must be a unique 4-digit number.", "Error", wx.ICON_ERROR)
        dialog.Destroy()

    def on_save_changes(self, event):
        file_path = os.path.join(app_dir, "listofaccountstoread.txt")
        self.save_accounts(file_path)
        wx.MessageBox(f"Changes saved successfully in {file_path}!", "Info", wx.ICON_INFORMATION)

    def on_remove_selected(self, event):
        """Remove the selected account."""
        selection = self.account_list.GetSelection()
        if selection != wx.NOT_FOUND:  # Ensure something is selected
            account = self.account_list.GetString(selection)
            self.accounts.remove(account)
            self.account_list.Delete(selection)
        else:
            wx.MessageBox("Please select an account to remove.", "Error", wx.ICON_WARNING)
    
    #def on_cancelation(self, event):
        #"""Close the dialog without saving"""
        #self.EndModal(wx.ID_CANCEL)

class MyApp(wx.App):
    def OnInit(self):
        frame = AccountManager(None, title="Account Manager")
        frame.Show()
        return True

class MyFrame(wx.Frame):
    def __init__(self, parent, title):
        super().__init__(parent, title=title, size=(1250, 750))
        
        self.panel = wx.Panel(self)
        self.sizer = wx.BoxSizer(wx.VERTICAL)

        # Create a horizontal sizer for the "Log In" button and the new button
        login_sizer = wx.BoxSizer(wx.HORIZONTAL)

        self.login_button = wx.Button(self.panel, label="Log In")
        self.ch_cred_button = wx.Button(self.panel, label="Change Credentials")
        self.acct_mngmt = wx.Button(self.panel, label="Manage Accounts")

        # Add the login button and the new button to the horizontal sizer
        login_sizer.Add(self.login_button, flag=wx.RIGHT, border=10)
        login_sizer.Add(self.ch_cred_button, flag=wx.LEFT, border=10)
        login_sizer.Add(self.acct_mngmt, flag=wx.LEFT, border=10)

        date_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.checkbox = wx.CheckBox(self.panel, label="Using saved accounts", pos=(50, 50))
        self.checkbox2 = wx.CheckBox(self.panel, label="Today", pos=(50, 50))
        
        self.label_start_date = wx.StaticText(self.panel, label="Start Date:")
        self.start_date_picker = wx.adv.DatePickerCtrl(self.panel, style=wx.adv.DP_DROPDOWN)

        self.label_end_date = wx.StaticText(self.panel, label="End Date:")
        self.end_date_picker = wx.adv.DatePickerCtrl(self.panel, style=wx.adv.DP_DROPDOWN)

        self.save_button = wx.Button(self.panel, label="Save Dates")
        
        date_sizer.Add(self.checkbox, flag=wx.ALL, border=10)
        date_sizer.Add(self.checkbox2, flag=wx.ALL, border=10)
        date_sizer.Add(self.label_start_date, flag=wx.ALL, border=10)
        date_sizer.Add(self.start_date_picker, flag=wx.ALL | wx.EXPAND, border=10)
        date_sizer.Add(self.label_end_date, flag=wx.ALL, border=10)
        date_sizer.Add(self.end_date_picker, flag=wx.ALL | wx.EXPAND, border=10)
        date_sizer.Add(self.save_button, flag=wx.ALL | wx.ALIGN_CENTER, border=10)        

        # Add the horizontal sizer with the login buttons to the main vertical sizer
        self.sizer.Add(login_sizer, flag=wx.ALL, border=10)
        self.sizer.Add(date_sizer, flag=wx.ALL, border=10)
        
        self.midrowsizer = wx.BoxSizer(wx.HORIZONTAL)
        self.check_button = wx.Button(self.panel, label="Get Operations")
        self.open_location = wx.Button(self.panel, label="File Location")
        self.midrowsizer.Add(self.check_button, flag=wx.ALL, border=10)
        self.midrowsizer.Add(self.open_location, flag=wx.ALL, border=10)

        # Continue adding the other elements to the main vertical sizer                                                     
        self.output_text_ctrl = wx.TextCtrl(self.panel, size=(300, 110), style=wx.TE_MULTILINE | wx.TE_READONLY)
        self.run_button = wx.Button(self.panel, label="View Operations")
           
        #self.sizer.Add()   
        self.sizer.Add(self.midrowsizer, flag=wx.ALL, border=10)                          
        self.sizer.Add(self.output_text_ctrl, flag=wx.ALL | wx.EXPAND, border=10)
        self.sizer.Add(self.run_button, flag=wx.ALL, border=10)

        self.panel.SetSizer(self.sizer)
        
        #self.check_button.Disable()
        self.checkbox.SetValue(True)
        self.checkbox2.SetValue(True)
        self.df2 = None
        self.df3 = None
        self.date_for_url = datetime.now().strftime(r"%d/%m/%Y")
        self.date_for_url2 = datetime.now().strftime(r"%d/%m/%Y")
        self.date_for_file = datetime.now().strftime("%d-%m-%Y")
        self.start_date_picker.Disable()
        self.end_date_picker.Disable()
        self.save_button.Disable()

        # Bind events
        self.login_button.Bind(wx.EVT_BUTTON, self.on_update_token)
        self.ch_cred_button.Bind(wx.EVT_BUTTON, self.on_modify_cred_file)
        self.acct_mngmt.Bind(wx.EVT_BUTTON, self.on_manage_accounts)
        self.check_button.Bind(wx.EVT_BUTTON, self.fetch_operations)
        self.open_location.Bind(wx.EVT_BUTTON, self.open_program_location)
        self.checkbox.Bind(wx.EVT_CHECKBOX, self.on_checkbox_toggled)
        self.checkbox2.Bind(wx.EVT_CHECKBOX, self.on_checkbox2_toggled)
        self.save_button.Bind(wx.EVT_BUTTON, self.save_dates)
        self.run_button.Bind(wx.EVT_BUTTON, self.run_function)
        
        self.df = None  # Dataframe placeholder
        
        # Redirect print statements to output_text_ctrl
        sys.stdout = TextRedirect(self.output_text_ctrl)

        self.Show()
    
    def on_checkbox_toggled(self, event):
        self.check_button.Enable(self.checkbox.IsChecked())
    
    def on_checkbox2_toggled(self, event):
        isnt_checked = not self.checkbox2.IsChecked()
        is_checked = self.checkbox2.IsChecked()
        #print(f"Checkbox state: {'Checked' if isnt_checked else 'Unchecked'}")
        self.start_date_picker.Enable(isnt_checked)
        self.end_date_picker.Enable(isnt_checked)
        self.save_button.Enable(isnt_checked)
        if is_checked:
            self.date_for_url = datetime.now().strftime("%d/%m/%Y")
            self.date_for_url2 = datetime.now().strftime("%d/%m/%Y")
            self.date_for_file = datetime.now().strftime("%d-%m-%Y")
            print(f"Fechas de búsqueda reseteadas al día de hoy: {self.date_for_url}")

    def save_dates(self, event):
        """Save the selected dates in the required format."""
        # Get the selected dates
        start_date = self.start_date_picker.GetValue()
        end_date = self.end_date_picker.GetValue()

        # Convert wx.DateTime to Python datetime.date
        start_date_py = datetime.strptime(start_date.FormatISODate(), "%Y-%m-%d")
        end_date_py = datetime.strptime(end_date.FormatISODate(), "%Y-%m-%d")

        # Format dates for the API as dd/mm/yyyy
        self.date_for_url = start_date_py.strftime(f"%d/%m/%Y")
        self.date_for_url2 = end_date_py.strftime(f"%d/%m/%Y")
        self.date_for_file = end_date_py.strftime("%d-%m-%Y")

        # Print or save the dates
        print(f"Fecha de comienzo definida en: {self.date_for_url}")
        print(f"Fecha de final definida en: {self.date_for_url2}")

        # You can save these dates to a file or use them elsewhere in your application
        file_path = os.path.join(app_dir, "dates.txt")  # Construct the full path

    # Save the dates to the file
        with open(file_path, "w", encoding="utf-8") as file:
            file.write(f"{self.date_for_url}\n")
            file.write(f"{self.date_for_url2}\n")

        wx.MessageBox("Dates saved successfully!", "Info", wx.ICON_INFORMATION)

    def on_modify_cred_file(self, event):
        """Allows to change login credentials"""
        dialog = UpdateCredentialsDialog(self)
        dialog.ShowModal()
        dialog.Destroy()

    def on_manage_accounts(self, event):
        """Allows to change login credentials"""
        dialog = AccountManager(self)
        dialog.ShowModal()
        dialog.Destroy()
    
    def on_update_token(self, event):
        with open(credentials_path, 'r') as file:
            creds = json.load(file)
        username = creds.get('username')
        password = creds.get('password')
        #print(creds)

        """Run API login and update token"""
        base_login_url = 'https://conosur.aunesa.com/Irmo/api/login' 
        loginheaders = {
            "Content-Type": "application/json"
        }
        loginbody = {
                "clientId": "conosur",   
                "username": username,  
                "password": password
            }
        response = requests.post(base_login_url, json=loginbody, headers=loginheaders)

        if response.status_code == 200 or response.status_code == 201:
            with open(authkey_path, "w") as file:
                json.dump(response.json(), file)
            print("Nuevo token guardado en authkey.json. El mismo tendrá validez por las próximas 24hs")                         
        else:
            print(f"Request failed with status code {response.status_code}: {response.text}")
    
    def open_program_location(self, event):
        """Open the directory of the running program."""
        try:
            if os.name == 'nt':  # Windows
                subprocess.Popen(f'explorer "{app_dir}"')
            elif os.name == 'posix':  # macOS or Linux
                subprocess.Popen(['open', app_dir]) if 'darwin' in os.sys.platform else subprocess.Popen(['xdg-open', app_dir])
        except Exception as e:
            wx.MessageBox(f"Could not open program location: {e}", "Error", wx.ICON_ERROR)

    #def run_function(self, event):
    #    print('This will show grids for the files it created')

    def fetch_operations(self, event):
        file_path = os.path.join(app_dir, "listofaccountstoread.txt")
        with open(authkey_path, "r") as file:
            authkey = json.load(file)
        token = list(authkey.values())[1]
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {token}"
        }       
        with open(file_path, "r") as file:
            accounts = file.read().splitlines()
        responses = []
        for account in accounts:
            uri = f"https://conosur.aunesa.com/Irmo/api/operaciones/informes?cuenta={account}&fechaDesde={self.date_for_url}&fechaHasta={self.date_for_url2}"    
            try:
                response = requests.get(uri, headers=headers)
                response.raise_for_status()  # Raise an exception for HTTP errors
                responses.append(response.json())
            except requests.exceptions.RequestException as e:
                print(f"Error recuperando operaciones para la cuenta {account}: {e}")
        final_json = json.dumps(responses, indent=4)
        file_name = f"InformedeOperaciones_{self.date_for_file}.json"
        file_path = os.path.join(app_dir, file_name)
        with open(file_path, "w", encoding="utf-8") as json_file:
            json_file.write(final_json)
        print(f"Data de operaciones guardada en {file_name}")


        today = datetime.now().strftime('%Y-%m-%d')
        json_file_path = os.path.join(app_dir, f'InformedeOperaciones_{self.date_for_file}.json')   #
        #json_file_path = f"C:\\Users\\sebastian.alfonso\\InformedeOperaciones_2024-11-21.json"

        with open(json_file_path, encoding='utf-8') as file:
            raw_data = json.load(file)
        data = raw_data[0]
        df = pd.DataFrame(data)     # DF IS THE RAW JSON RECEIVED FROM AUNE'S API IN DATAFRAME FORM
        #print("Columns:", df.columns)
        #print(df.head())
        df['instrumento'] = df['instrumento'].astype(str).str.strip()
        dfb = df[df['instrumento'].str.match(r'^\[DLR\d{6}\]$', na=False)] # DFB IS JUST ROFEX OPERATIONS
        df = df[~df['instrumento'].str.match(r'^\[DLR\d{6}\]$', na=False)]
        df2 = pd.DataFrame()    # DF2 RESULTS IN THE PROCESSED LIST OF "BYMA" OPERATIONS.
        df3 = pd.DataFrame()    # DF3 WILL PROCESS ROFEX OPERATIONS

        def clean_currency(entry):
            if isinstance(entry, list) and len(entry) > 0 and isinstance(entry[0], dict):
                moneda = entry[0].get('Moneda', '')
                monto = entry[0].get('Monto', '')

                if not monto or monto in ['0', '0.0', '']:
                    return 0
                return f"{moneda} {monto}"
            return 0

        def isthiscompra(row):
            if pd.isna(row):
                return None
            if "Concurrencia Contado - Compra" in row:
                return 'Compra'
            elif "SENEBI Contado - Compra" in row:
                return 'Compra'
            elif "Futuros Financieros - Compra" in row:
                return 'Compra'
            elif "Futuros Financieros - Venta" in row:
                return 'Venta'
            elif "SENEBI Contado - Venta" in row:
                return 'Venta'
            elif "Concurrencia Contado - Venta" in row:
                return 'Venta'
            return None

        def whattasa(row):
            if "ARS 24hs" in row:
                return '1'
            elif "USD 24hs" in row:
                return '1'
            elif "ARS 24hs NG" in row:
                return '1'
            elif "USD 24hs NG" in row:
                return '1'
            elif "USDC 24hs NG" in row:
                return '1'
            elif 'ARS Inm' in row:
                return '0'
            elif 'USD Inm' in row:
                return '0'
            elif 'USD' in row:
                return '0'
            elif 'ARS' in row:
                return '0'
    
        def whattasa2(row):
            if "Futuros Financieros - Venta" in row:
                return '1'
            elif "Futuros Financieros - Compra" in row:
                return '1'
            else:
                return '1'

        def update_mercado(df, df2):
            condition = df['tipoOperacion'].isin(["SENEBI Contado - Compra", "SENEBI Contado - Venta"])
            df2['Mercado'] = condition.map({True: "BYMA SENEBI", False: "BYMA"})
            return df2

        def extract_monto(value):
            if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                monto = value[0].get('Monto', '0')
                if isinstance(monto, str):
                    monto = monto.replace(',', '.')
                return float(monto)
            return 0

        temp_results = df['neto'].apply(clean_currency)
        df['Neto'] = temp_results
        df['Monedas'] = df['Neto'].apply(lambda x: x.split()[0] if isinstance(x, str) and len(x.split()) > 1 else '')

        def pesodolar(row):
            if "USD" in row:
                return 'U$D'
            elif "ARS" in row:
                return '$'

        # PROCESSING BYMAS
        if not df.empty:
            df['concertacion'] = pd.to_datetime(df['concertacion'], errors='coerce')
            df['liquidacion'] = pd.to_datetime(df['liquidacion'], errors='coerce')                                   
            df2['Fecha Concer.'] = df['concertacion'].dt.strftime('%d/%m/%Y')
            df2['F. Liqui'] = df['liquidacion'].dt.strftime('%d/%m/%Y')
            df2['FONDO'] = '32STGESTION VII'
            df2['OPERACION'] = df['tipoOperacion'].apply(isthiscompra)
            df2['Especie'] = df['instrumento'].str.extract(r'\] (\w+)')
            df2['Contraparte'] = 'Cono Sur Inversiones S.A.'
            df2['Plazo Liq.'] = df['condiciones'].apply(whattasa)
            df2['TASA'] = ''
            df2['Valor Nominal'] = df['cantidadTotal']
            df2['Precio'] = df['precioPromedio'].abs()
            df2['Precio'] = df2['Precio'].round(2)
            df2['Importe'] = df['bruto']
            df2['Importe'] = df2['Importe'].abs()
            df2['Moneda'] = df['Monedas'].apply(pesodolar)
            df2 = update_mercado(df, df2)
            def adjust_importe(df2):
                df2.loc[df2['Mercado'] == "BYMA SENEBI", 'Importe'] /= 100
                return df2
            df2 = adjust_importe(df2)
            df2['GASTOS'] = '0'
            file_name = f"operacionesbyma_{today}.xlsx"  # Construct the filename
            file_path = os.path.join(app_dir, file_name)  # Construct the full path           
            df2.to_excel(file_path, index=False)
            print('Fetching BYMA operations...')
            #print(df2)
            print(f"BYMA operationas saved to {self.date_for_file}.xlsx")
            wb = load_workbook(file_path)
            ws = wb.active
            green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
            for cell in ws[1]:
                cell.fill = green_fill
            wb.save(file_path)
            #print(f"Excel file saved with styled headers at {excel_file_path2}")
            print('...')
        else:
            print("No BYMA operations. Skipping processing.")
        
        self.df2 = pd.DataFrame()
        self.df3 = pd.DataFrame()
        self.df2 = df2
        self.df3= df3

        if not dfb.empty and len(dfb) > 0:
            temp_results_b = dfb['neto'].apply(clean_currency)
            #print('neto')
            #print(dfb['neto'])
            dfb['Neto'] = temp_results_b
            dfb['Monedas'] = dfb['Neto'].apply(lambda x: x.split()[0] if isinstance(x, str) and len(x.split()) > 1 else '')
            dfb['concertacion'] = pd.to_datetime(dfb['concertacion'], errors='coerce')
            dfb['liquidacion'] = pd.to_datetime(dfb['liquidacion'], errors='coerce') 
            df3['Fecha Concer.'] = dfb['concertacion'].dt.strftime('%d/%m/%Y')
            df3['F. Liqui'] = dfb['liquidacion'].dt.strftime('%d/%m/%Y')
            df3['FONDO'] = '32STGESTION VII'
            df3['OPERACION'] = dfb['tipoOperacion'].apply(isthiscompra)
            df3['Especie'] = dfb['instrumento']
            df3['Contraparte'] = 'Cono Sur Inversiones S.A.'
            df3['Plazo Liq.'] = dfb['tipoOperacion'].apply(whattasa2)
            df3['TASA'] = ''
            df3['Valor Nominal'] = dfb['cantidadTotal']
            df3['Precio'] = dfb['precioPromedio'].abs()
            df3['Precio'] = df3['Precio'].round(2)
            df3['Importe'] = dfb['neto'].apply(extract_monto)
            df3['Importe'] = pd.to_numeric(df3['Importe'], errors='coerce').fillna(0)
            df3['Importe'] = df3['Importe'].abs()
            df3['Moneda'] = dfb['Monedas'].apply(pesodolar)
            df3['Mercado'] = 'ROFEX'
            dfb['Gasto'] = dfb['gastos'].apply(extract_monto)
            dfb['Impuesto'] = dfb['impuestos'].apply(extract_monto)
            dfb['Gastos'] = pd.to_numeric(dfb['Gasto'], errors='coerce').fillna(0)
            dfb['Impuestos'] = pd.to_numeric(dfb['Impuesto'], errors='coerce').fillna(0)
            df3['GASTOS'] = dfb['Gastos'] + dfb['Impuestos']
            file_name3 = f'operaciones_rofex_{self.date_for_file}.xlsx' #ROFEX
            file_path3 = os.path.join(app_dir, file_name3)
            df3.to_excel(file_path3, index=False)
            print('Fetching ROFEX operations...')
            #print(df3)
            print(f"ROFEX operationas saved to {file_path3}")
            wb = load_workbook(file_path3)
            ws = wb.active
            green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
            for cell in ws[1]:
                cell.fill = green_fill
            wb.save(file_path3)
            #print(f"Excel file saved with styled headers at {excel_file_path3}")
        else:
            print("No ROFEX operations. Skipping processing.")

        self.df2 = pd.DataFrame()
        self.df3 = pd.DataFrame()

        if not df2.empty:                   
            self.df2 = df2
        
        if not df2.empty:
            self.df3= df3            

    def run_function(self, event):
    # Safely check for the existence of DataFrames
        if hasattr(self, 'df2') and hasattr(self, 'df3') and (self.df2 is not None or self.df3 is not None):
            print("Mostrando grillas de operaciones")

            # Hide old grids if they exist
            if hasattr(self, 'grid_sizer') and self.grid_sizer is not None:
                self.grid_sizer.Clear(True)

            self.grid_sizer = wx.BoxSizer(wx.VERTICAL)

            # Create and populate the grid for df2
            self.grid = wx.grid.Grid(self.panel)
            self.grid.CreateGrid(len(self.df2), len(self.df2.columns))
            for col, header in enumerate(self.df2.columns):
                self.grid.SetColLabelValue(col, header)
            for row in range(len(self.df2)):
                for col in range(len(self.df2.columns)):
                    value = self.df2.iloc[row, col]
                    if col == self.df2.columns.get_loc('Importe'):  # Fixed to self.df2
                        formatted_value = f"{value:.2f}"
                        self.grid.SetCellValue(row, col, formatted_value)
                        self.grid.SetCellAlignment(row, col, wx.ALIGN_RIGHT, wx.ALIGN_CENTER)
                    else:
                        self.grid.SetCellValue(row, col, str(value))

            # Create and populate the grid for df3
            self.gris = wx.grid.Grid(self.panel)
            self.gris.CreateGrid(len(self.df3), len(self.df3.columns))
            for col, header in enumerate(self.df3.columns):
                self.gris.SetColLabelValue(col, header)
            for row in range(len(self.df3)):
                for col in range(len(self.df3.columns)):
                    value = self.df3.iloc[row, col]
                    if col == self.df3.columns.get_loc('Importe'):  # Fixed to self.df3
                        formatted_value = f"{value:.2f}"
                        self.gris.SetCellValue(row, col, formatted_value)
                        self.gris.SetCellAlignment(row, col, wx.ALIGN_RIGHT, wx.ALIGN_CENTER)
                    else:
                        self.gris.SetCellValue(row, col, str(value))

            # Add the grids to the grid_sizer
            self.grid_sizer.Add(self.grid, flag=wx.EXPAND | wx.ALL, border=10)
            self.grid_sizer.Add(self.gris, flag=wx.EXPAND | wx.ALL, border=10)

            # Add the updated grid_sizer to the main sizer
            self.sizer.Add(self.grid_sizer, flag=wx.EXPAND)

            # Refresh the layout
            self.panel.Layout()
            self.panel.Fit()
            self.Layout()

            # Show the grids
            self.grid.Show()
            self.gris.Show()

        else:
            print("No hay operaciones cargadas aún")

if __name__ == "__main__":
    app = wx.App(False)
    frame = MyFrame(None, "[C.O.N.I.T.O.] Chequeador Óptimo de Nuevas Interacciones a Título Oneroso")
    app.MainLoop()   
